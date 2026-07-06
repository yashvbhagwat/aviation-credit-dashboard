import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import date
import io
import re
import statistics
import requests
import yfinance as yf

st.set_page_config(
    page_title="Aviation Finance Dashboard",
    page_icon="\u2708",
    layout="wide",
)

# RAG hex colors
GREEN_BG, GREEN_TEXT = "#f0fdf4", "#15803d"
AMBER_BG, AMBER_TEXT = "#fffbeb", "#b45309"
RED_BG, RED_TEXT = "#fef2f2", "#b91c1c"
NM_BG, NM_TEXT = "#f9fafb", "#6b7280"

COLOR_MAP = {
    "green": (GREEN_BG, GREEN_TEXT),
    "amber": (AMBER_BG, AMBER_TEXT),
    "red": (RED_BG, RED_TEXT),
    "nm": (NM_BG, NM_TEXT),
}

AIRLINE_SUGGESTIONS = {
    "Delta Air Lines": "DAL", "United Airlines Holdings": "UAL",
    "American Airlines Group": "AAL", "Southwest Airlines": "LUV",
    "Alaska Air Group": "ALK", "JetBlue Airways": "JBLU",
    "Allegiant Travel Company": "ALGT", "Frontier Group Holdings": "ULCC",
    "Spirit Airlines": "SAVE", "Sun Country Airlines": "SNCY",
    "Hawaiian Holdings": "HA", "SkyWest Inc": "SKYW",
    "Air Transport Services Group": "ATSG", "Mesa Air Group": "MESA",
    "Lufthansa Group": "LHA.DE", "Ryanair Holdings": "RYAAY",
    "International Airlines Group (IAG)": "IAG.L", "Air France-KLM": "AF.PA",
    "easyJet": "EZJ.L", "Wizz Air Holdings": "WIZZ.L",
    "Singapore Airlines": "C6L.SI", "Cathay Pacific Airways": "0293.HK",
    "Air China": "0753.HK", "China Southern Airlines": "1055.HK",
    "China Eastern Airlines": "0670.HK", "Japan Airlines": "9201.T",
    "ANA Holdings": "9202.T", "Qantas Airways": "QAN.AX",
    "Korean Air Lines": "003490.KS", "Turkish Airlines": "THYAO.IS",
    "Air Canada": "AC.TO", "Copa Holdings": "CPA", "LATAM Airlines": "LTM",
    "Gol Linhas Aereas": "GOLL4.SA", "Azul": "AZUL",
    "IndiGo (InterGlobe Aviation)": "INDIGO.NS", "SpiceJet": "SPICEJET.NS",
}

# id, label, key, unit_symbol, dimension
RATIOS = [
    (1, "Current Ratio", "current_ratio", "x", "LIQUIDITY"),
    (2, "Quick Ratio", "quick_ratio", "x", "LIQUIDITY"),
    (3, "Cash % Revenue", "cash_pct_revenue", "%", "LIQUIDITY"),
    (4, "Net Debt / EBITDAR", "net_debt_ebitdar", "x", "LEVERAGE"),
    (5, "Total Debt / Assets", "debt_assets", "x", "LEVERAGE"),
    (6, "Total Debt / Equity", "debt_equity", "x", "LEVERAGE"),
    (7, "Interest Coverage", "interest_coverage", "x", "COVERAGE"),
    (8, "EBITDAR Coverage", "ebitdar_coverage", "x", "COVERAGE"),
]
RATIO_BY_ID = {r[0]: r for r in RATIOS}
HIGHER_BETTER = {1, 2, 3, 7, 8}
LEVERAGE_IDS = {4, 5, 6}

ABS_THRESHOLDS = {
    # id: (green_cut, amber_cut). For higher-better: v>=green green, v>=amber amber, else red.
    # For lower-better: v<=green green, v<=amber amber, else red.
    1: (0.90, 0.60),
    2: (0.80, 0.50),
    3: (15.0, 10.0),
    4: (4.0, 6.0),
    5: (0.70, 0.85),
    6: (3.0, 5.0),
    7: (2.0, 1.0),
    8: (3.0, 2.0),
}
# threshold used for chart reference lines (the green cutoff)
GREEN_LINE = {rid: ABS_THRESHOLDS[rid][0] for rid in ABS_THRESHOLDS}

def safe_div(a, b):
    if a is None or b is None or b == 0:
        return None
    return a / b

# ----------------------------------------------------------------------------
# RAG colour resolution
# ----------------------------------------------------------------------------
def absolute_rag(rid, v):
    if v is None:
        return None
    g, a = ABS_THRESHOLDS[rid]
    if rid in HIGHER_BETTER:
        if v >= g:
            return "green"
        if v >= a:
            return "amber"
        return "red"
    if v <= g:
        return "green"
    if v <= a:
        return "amber"
    return "red"


def peer_rag(rid, v, peers):
    vals = [x for x in peers if x is not None]
    if v is None or len(vals) < 2:
        return absolute_rag(rid, v)
    med = statistics.median(vals)
    try:
        qs = statistics.quantiles(vals, n=4)
    except statistics.StatisticsError:
        qs = [med, med, med]
    q1, q3 = qs[0], qs[2]
    if rid in LEVERAGE_IDS:
        if v <= med:
            return "green"
        if v >= q3:
            return "red"
        return "amber"
    if v >= med:
        return "green"
    if v <= q1:
        return "red"
    return "amber"


def base_color(rid, v, peers, peer_mode):
    if peer_mode:
        return peer_rag(rid, v, peers)
    return absolute_rag(rid, v)


def resolve_color(rid, records, peers, peer_mode):
    """Latest-year colour name for a ratio, including N/M trend rules."""
    if not records:
        return "nm"
    latest = records[-1]
    v = latest["ratios"].get(rid)
    if v is not None:
        return base_color(rid, v, peers, peer_mode) or "nm"

    # Latest is None. Apply N/M trend logic for ratios 4 and 6, else neutral.
    prior = records[-2] if len(records) >= 2 else None
    if rid == 4 and latest["nm"].get(4):
        cur_e = latest["ebitdar"]
        prior_e = prior["ebitdar"] if prior else None
        if prior_e is None or prior_e <= 0:
            return "red"
        if cur_e is not None and cur_e > prior_e:
            return "amber"
        return "red"
    if rid == 6 and latest["nm"].get(6):
        cur_q = latest["equity"]
        prior_q = prior["equity"] if prior else None
        if prior_q is None:
            return "red"
        if cur_q is not None and cur_q > prior_q:
            return "amber"
        return "red"
    return "nm"


def fmt_ratio(rid, v):
    if v is None:
        return "N/M"
    sym = RATIO_BY_ID[rid][3]
    if sym == "%":
        return f"{v:.1f}%"
    return f"{v:.2f}x"


def to_millions(v):
    if v is None:
        return None
    return v / 1_000_000.0

def fmt_mcap(m):
    if m is None:
        return "\u2014"
    if m >= 1_000_000_000:
        return f"${m / 1e9:.1f}B"
    if m >= 1_000_000:
        return f"${m / 1e6:.1f}M"
    return f"${m:,.0f}"


def price_line_html(stock):
    price = stock.get("price")
    prev = stock.get("prev")
    mcap = stock.get("mcap")
    if price is None:
        return "<span style='color:#9ca3af;'>price unavailable</span>"
    chg_html = ""
    if prev:
        chg = price - prev
        pct = chg / prev * 100 if prev else 0.0
        arrow = "\u25b2" if chg >= 0 else "\u25bc"
        col = GREEN_TEXT if chg >= 0 else RED_TEXT
        chg_html = f" <span style='color:{col};'>{arrow} {pct:+.1f}%</span>"
    mcap_html = f" | Mkt Cap: {fmt_mcap(mcap)}" if mcap else ""
    return f"${price:,.2f}{chg_html}{mcap_html}"

# yfinance financial-statements engine (Financial Statements tab)
# ----------------------------------------------------------------------------
# Row spec: (yf_label, display_label, style, special, direction)
#   style:     "line" | "bold" (#f8fafc) | "sub" (#f1f5f9)
#   special:   None (=> $M, 2dp) | "eps" ($/share) | "pct" (%)
#   direction: "up" (green if rising) | "down" (red if rising) | None (neutral)
# yfinance row labels drift across versions; matching is case/space-insensitive
# with a small alias map, and missing rows are skipped silently per spec.

YF_ALIASES = {
    "operating expense": ["operating expenses", "total operating expenses"],
    "operating income": ["operating income loss", "ebit"],
    "tax provision": ["income tax expense", "tax provision benefit"],
    "pretax income": ["income before tax", "pre tax income"],
    "net ppe": ["net property plant and equipment", "property plant and equipment net"],
    "total liabilities net minority interest": ["total liabilities"],
    "total equity gross minority interest": ["total equity"],
    "stockholders equity": ["common stock equity", "total stockholders equity"],
    "operating cash flow": ["cash flow from operations", "total cash from operating activities"],
    "investing cash flow": ["total cash from investing activities"],
    "financing cash flow": ["total cash from financing activities"],
    "end cash position": ["ending cash position", "end cash position"],
    "total revenue": ["total revenues", "revenue"],
}

INCOME_YF = [
    ("Total Revenue", "Total Revenue", "line", None, "up"),
    ("Cost Of Revenue", "Cost of Revenue", "line", None, "down"),
    ("Gross Profit", "Gross Profit", "sub", None, "up"),
    ("Operating Expense", "Total Operating Expenses", "bold", None, "down"),
    ("Operating Income", "Operating Income / EBIT", "bold", None, "up"),
    ("EBITDA", "EBITDA", "bold", None, None),
    ("Interest Expense", "Interest Expense", "line", None, "down"),
    ("Interest Income", "Interest Income", "line", None, "up"),
    ("Pretax Income", "Income Before Tax", "line", None, "up"),
    ("Tax Provision", "Income Tax Expense", "line", None, "down"),
    ("Net Income", "Net Income", "bold", None, "up"),
    ("__NET_MARGIN__", "Net Margin %", "line", "pct", "up"),
    ("Basic EPS", "EPS Basic ($)", "eps", "eps", None),
    ("Diluted EPS", "EPS Diluted ($)", "eps", "eps", None),
    ("Normalized EBITDA", "Normalized EBITDA", "line", None, None),
]

BALANCE_YF = [
    ("Cash And Cash Equivalents", "Cash & Equivalents", "line", None, None),
    ("Short Term Investments", "Short-term Investments", "line", None, None),
    ("Accounts Receivable", "Accounts Receivable", "line", None, None),
    ("Inventory", "Inventory", "line", None, None),
    ("Current Assets", "Total Current Assets", "sub", None, None),
    ("Net PPE", "Property, Plant & Equipment (net)", "line", None, None),
    ("Goodwill", "Goodwill", "line", None, None),
    ("Total Assets", "Total Assets", "sub", None, None),
    ("Accounts Payable", "Accounts Payable", "line", None, None),
    ("Current Debt", "Current Portion of Debt", "line", None, "down"),
    ("Current Liabilities", "Total Current Liabilities", "sub", None, "down"),
    ("Long Term Debt", "Long-term Debt", "line", None, "down"),
    ("Total Liabilities Net Minority Interest", "Total Liabilities", "sub", None, "down"),
    ("Stockholders Equity", "Total Stockholders Equity", "bold", None, "up"),
    ("Total Equity Gross Minority Interest", "Total Equity", "sub", None, "up"),
    ("Total Capitalization", "Total Capitalization", "line", None, None),
]

CASHFLOW_YF = [
    ("Net Income", "Net Income", "line", None, "up"),
    ("Depreciation And Amortization", "Depreciation & Amortization", "line", None, None),
    ("Change In Working Capital", "Changes in Working Capital", "line", None, None),
    ("Operating Cash Flow", "Cash from Operations", "sub", None, None),
    ("Capital Expenditure", "Capital Expenditures", "line", None, None),
    ("Purchase Of Investment", "Purchases of Investments", "line", None, None),
    ("Sale Of Investment", "Proceeds from Investment Sales", "line", None, None),
    ("Investing Cash Flow", "Cash from Investing", "sub", None, None),
    ("Issuance Of Debt", "Debt Proceeds", "line", None, None),
    ("Repayment Of Debt", "Debt Repayments", "line", None, "down"),
    ("Repurchase Of Capital Stock", "Share Repurchases", "line", None, None),
    ("Payment Of Dividends", "Dividends Paid", "line", None, None),
    ("Financing Cash Flow", "Cash from Financing", "sub", None, None),
    ("End Cash Position", "Ending Cash Balance", "bold", None, None),
    ("Free Cash Flow", "Free Cash Flow", "bold", None, "up"),
    ("__FCF_MARGIN__", "FCF Margin %", "line", "pct", "up"),
]

YF_BG = {"bold": "#f8fafc", "sub": "#f1f5f9", "line": "transparent"}

def _yf_find(df, yf_label):
    """Resolve a display/yf label to an actual DataFrame index label, or None."""
    if df is None or df.empty:
        return None
    norm = {str(idx).strip().lower(): idx for idx in df.index}
    candidates = [yf_label.strip().lower()]
    candidates += YF_ALIASES.get(yf_label.strip().lower(), [])
    for c in candidates:
        if c in norm:
            return norm[c]
    return None


def _yf_series(df, yf_label, cols):
    idx = _yf_find(df, yf_label)
    if idx is None:
        return {c: None for c in cols}
    row = df.loc[idx]
    out = {}
    for c in cols:
        try:
            v = row[c]
            out[c] = None if pd.isna(v) else float(v)
        except Exception:
            out[c] = None
    return out


def _col_label(ts, period):
    try:
        if period == "Annual":
            return f"FY{ts.year}"
        return f"Q{(ts.month - 1) // 3 + 1} {ts.year}"
    except Exception:
        return str(ts)


def prepare_yf_statement(df, spec, period, n, revenue_df=None, fcf_df=None):
    """Return (col_labels_ascending, columns, rows). Rows include calc rows."""
    if df is None or df.empty:
        return [], [], []
    cols = list(df.columns)[:n]            # yfinance: most recent first
    cols = sorted(cols)                    # ascending for display
    labels = [_col_label(c, period) for c in cols]

    # revenue per column for margins (match by column label)
    rev_by_label = {}
    if revenue_df is not None and not revenue_df.empty:
        rcols = sorted(list(revenue_df.columns)[:n])
        rser = _yf_series(revenue_df, "Total Revenue", rcols)
        rev_by_label = {_col_label(c, period): rser.get(c) for c in rcols}

    rows = []
    for yf_label, disp, style, special, direction in spec:
        if yf_label == "__NET_MARGIN__":
            ni = next((r for r in rows if r["disp"] == "Net Income"), None)
            vals = {}
            for c in cols:
                lab = _col_label(c, period)
                rev = rev_by_label.get(lab)
                niv = ni["raw"].get(c) if ni else None
                vals[c] = (niv / rev * 100) if (ni and niv is not None and rev not in (None, 0)) else None
            rows.append({"disp": disp, "style": style, "special": special,
                         "dir": direction, "raw": vals})
            continue
        if yf_label == "__FCF_MARGIN__":
            fcf = next((r for r in rows if r["disp"] == "Free Cash Flow"), None)
            vals = {}
            for c in cols:
                lab = _col_label(c, period)
                rev = rev_by_label.get(lab)
                fv = fcf["raw"].get(c) if fcf else None
                vals[c] = (fv / rev * 100) if (fcf and fv is not None and rev not in (None, 0)) else None
            rows.append({"disp": disp, "style": style, "special": special,
                         "dir": direction, "raw": vals})
            continue
        series = _yf_series(df, yf_label, cols)
        if all(v is None for v in series.values()):
            continue
        rows.append({"disp": disp, "style": style, "special": special,
                     "dir": direction, "raw": series})
    return labels, cols, rows


def _yf_fmt(v, special):
    if v is None:
        return "\u2014", False
    if special == "eps":
        return f"${v:,.2f}", v < 0
    if special == "pct":
        return f"{v:.1f}%", v < 0
    return f"{v / 1e6:,.2f}", v < 0


def _yf_pct(raw, cols, direction):
    if len(cols) < 2:
        return "\u2014", NM_TEXT
    a, b = raw.get(cols[-2]), raw.get(cols[-1])
    if a is None or a == 0 or b is None:
        return "\u2014", NM_TEXT
    p = (b - a) / abs(a)
    txt = f"{p * 100:+.1f}%"
    if direction == "up":
        col = GREEN_TEXT if p > 0 else (RED_TEXT if p < 0 else NM_TEXT)
    elif direction == "down":
        col = RED_TEXT if p > 0 else (GREEN_TEXT if p < 0 else NM_TEXT)
    else:
        col = NM_TEXT
    return txt, col


def render_yf_html(title, labels, cols, rows, period, balance_extra=None):
    head = "".join(f"<th style='text-align:right;padding:6px 10px;'>{h}</th>" for h in labels)
    pct_hdr = f"% \u0394 ({labels[-2]}\u2192{labels[-1]})" if len(labels) >= 2 else "% \u0394"
    html = [
        f"<div style='font-weight:700;font-size:16px;margin:14px 0 4px;'>{title}</div>",
        "<table style='width:100%;border-collapse:collapse;font-size:13px;'>",
        f"<tr style='border-bottom:2px solid #cbd5e1;'>"
        f"<th style='text-align:left;padding:6px 10px;'>Line Item</th>{head}"
        f"<th style='text-align:right;padding:6px 10px;'>{pct_hdr}</th></tr>",
    ]
    for r in rows:
        bg = YF_BG.get(r["style"], "transparent")
        weight = "700" if r["style"] in ("bold", "sub") else "400"
        cells = ""
        for c in cols:
            txt, isneg = _yf_fmt(r["raw"].get(c), r["special"])
            color = RED_TEXT if isneg else "#111827"
            cells += f"<td style='text-align:right;padding:5px 10px;color:{color};'>{txt}</td>"
        ptxt, pcol = _yf_pct(r["raw"], cols, r["dir"])
        html.append(
            f"<tr style='background:{bg};border-bottom:1px solid #f1f5f9;'>"
            f"<td style='padding:5px 10px;font-weight:{weight};'>{r['disp']}</td>{cells}"
            f"<td style='text-align:right;padding:5px 10px;color:{pcol};font-weight:{weight};'>{ptxt}</td></tr>"
        )
    if balance_extra is not None:
        assets, tl, te = balance_extra
        chk = ""
        for c in cols:
            a = assets.get(c)
            le = (tl.get(c) or 0) + (te.get(c) or 0) if (tl.get(c) is not None or te.get(c) is not None) else None
            if a is None or le is None or a == 0:
                chk += "<td style='text-align:right;padding:5px 10px;color:#6b7280;'>\u2014</td>"
            elif abs(a - le) < 0.01 * abs(a):
                chk += f"<td style='text-align:right;padding:5px 10px;color:{GREEN_TEXT};font-weight:700;'>\u2713 Balanced</td>"
            else:
                chk += f"<td style='text-align:right;padding:5px 10px;color:{AMBER_TEXT};font-weight:700;'>\u26a0 Review</td>"
        html.append(f"<tr style='background:{YF_BG['sub']};'>"
                    f"<td style='padding:5px 10px;font-weight:700;'>Balance Check</td>{chk}"
                    f"<td style='padding:5px 10px;'></td></tr>")
    html.append("</table>")
    return "".join(html)


def _xl_write_yf(ws, labels, cols, rows, balance_extra=None):
    bold = Font(bold=True)
    red = Font(color="b91c1c")
    red_bold = Font(bold=True, color="b91c1c")
    sub_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    bold_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    header = ["Line Item"] + labels + ([f"% Change {labels[-2]}->{labels[-1]}"] if len(labels) >= 2 else ["% Change"])
    for j, h in enumerate(header, start=1):
        ws.cell(row=1, column=j, value=h).font = bold
    rn = 2
    for r in rows:
        is_b = r["style"] in ("bold", "sub")
        ws.cell(row=rn, column=1, value=r["disp"]).font = bold if is_b else None
        for k, c in enumerate(cols, start=2):
            v = r["raw"].get(c)
            if v is None:
                ws.cell(row=rn, column=k, value=None)
                continue
            if r["special"] == "eps":
                out = round(v, 2)
            elif r["special"] == "pct":
                out = round(v, 1)
            else:
                out = round(v / 1e6, 2)
            cell = ws.cell(row=rn, column=k, value=out)
            if out < 0:
                cell.font = red_bold if is_b else red
            elif is_b:
                cell.font = bold
        ptxt, _ = _yf_pct(r["raw"], cols, r["dir"])
        ws.cell(row=rn, column=len(cols) + 2, value=(ptxt if ptxt != "\u2014" else None))
        if is_b:
            fill = sub_fill if r["style"] == "sub" else bold_fill
            for col in range(1, len(cols) + 3):
                ws.cell(row=rn, column=col).fill = fill
        rn += 1
    if balance_extra is not None:
        assets, tl, te = balance_extra
        ws.cell(row=rn, column=1, value="Balance Check").font = bold
        for k, c in enumerate(cols, start=2):
            a = assets.get(c)
            le = (tl.get(c) or 0) + (te.get(c) or 0) if (tl.get(c) is not None or te.get(c) is not None) else None
            txt = "\u2014" if (a is None or le is None or a == 0) else ("Balanced" if abs(a - le) < 0.01 * abs(a) else "Review")
            ws.cell(row=rn, column=k, value=txt).font = bold
    ws.column_dimensions["A"].width = 34
    for k in range(2, len(cols) + 3):
        ws.column_dimensions[ws.cell(row=1, column=k).column_letter].width = 15


def build_yf_excel(ticker, stmts):
    wb = Workbook()
    plans = [
        ("Income Statement - Annual", stmts["fin"], INCOME_YF, "Annual", 3),
        ("Income Statement - Quarterly", stmts["qfin"], INCOME_YF, "Quarterly", 4),
        ("Balance Sheet - Annual", stmts["bs"], BALANCE_YF, "Annual", 3),
        ("Balance Sheet - Quarterly", stmts["qbs"], BALANCE_YF, "Quarterly", 4),
        ("Cash Flow - Annual", stmts["cf"], CASHFLOW_YF, "Annual", 3),
        ("Cash Flow - Quarterly", stmts["qcf"], CASHFLOW_YF, "Quarterly", 4),
    ]
    first = True
    for sheet_name, df, spec, period, n in plans:
        ws = wb.active if first else wb.create_sheet(sheet_name)
        if first:
            ws.title = sheet_name
            first = False
        rev_df = stmts["fin"] if period == "Annual" else stmts["qfin"]
        labels, cols, rows = prepare_yf_statement(df, spec, period, n, revenue_df=rev_df)
        if not labels:
            ws.cell(row=1, column=1, value="No data available from Yahoo Finance.")
            continue
        extra = None
        if "Balance" in sheet_name:
            extra = (_yf_series(df, "Total Assets", cols),
                     _yf_series(df, "Total Liabilities Net Minority Interest", cols),
                     _yf_series(df, "Total Equity Gross Minority Interest", cols))
        _xl_write_yf(ws, labels, cols, rows, balance_extra=extra)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ----------------------------------------------------------------------------
# yfinance data layer (single source for all tabs)
# ----------------------------------------------------------------------------
YF_ALIASES.update({
    "long term debt": ["long term debt and capital lease obligation"],
    "current debt": ["current debt and capital lease obligation",
                     "current portion of long term debt"],
    "current assets": ["total current assets"],
    "current liabilities": ["total current liabilities"],
    "operating lease liability": ["operating lease liabilities",
                                  "long term capital lease obligation"],
    "capital expenditure": ["capital expenditures", "purchase of ppe"],
    "stockholders equity": ["common stock equity", "total stockholders equity",
                            "total equity gross minority interest"],
})


@st.cache_data(show_spinner=False, ttl=900)
def load_ticker_bundle(ticker):
    b = {"info": {}, "fin": pd.DataFrame(), "bs": pd.DataFrame(), "cf": pd.DataFrame(),
         "qfin": pd.DataFrame(), "qbs": pd.DataFrame(), "qcf": pd.DataFrame(),
         "history": pd.DataFrame()}
    try:
        t = yf.Ticker(ticker)
    except Exception:
        return b
    try:
        b["info"] = t.info or {}
    except Exception:
        b["info"] = {}
    getters = {
        "fin": lambda: t.financials, "bs": lambda: t.balance_sheet, "cf": lambda: t.cashflow,
        "qfin": lambda: t.quarterly_financials, "qbs": lambda: t.quarterly_balance_sheet,
        "qcf": lambda: t.quarterly_cashflow, "history": lambda: t.history(period="5y"),
    }
    for k, g in getters.items():
        try:
            df = g()
            b[k] = df if isinstance(df, pd.DataFrame) else pd.DataFrame()
        except Exception:
            b[k] = pd.DataFrame()
    return b


def _year_map(df, label):
    if df is None or getattr(df, "empty", True):
        return {}
    idx = _yf_find(df, label)
    if idx is None:
        return {}
    row = df.loc[idx]
    out = {}
    for col in df.columns:
        try:
            y = col.year
            v = row[col]
            out[y] = None if pd.isna(v) else float(v)
        except Exception:
            continue
    return out


def extract_yf_financials(bundle):
    inc, bal, cf = bundle.get("fin"), bundle.get("bs"), bundle.get("cf")
    return {
        "revenue": _year_map(inc, "Total Revenue"),
        "assets_current": _year_map(bal, "Current Assets"),
        "liabilities_current": _year_map(bal, "Current Liabilities"),
        "cash": _year_map(bal, "Cash And Cash Equivalents"),
        "sti": _year_map(bal, "Other Short Term Investments"),
        "assets_total": _year_map(bal, "Total Assets"),
        "debt_current": _year_map(bal, "Current Debt"),
        "debt_noncurrent": _year_map(bal, "Long Term Debt"),
        "lease_total": _year_map(bal, "Operating Lease Liability"),
        "equity": _year_map(bal, "Stockholders Equity"),
        "operating_income": _year_map(inc, "Operating Income"),
        "interest_expense": _year_map(inc, "Interest Expense"),
        "da": _year_map(cf, "Depreciation And Amortization"),
        "op_lease_cost": _year_map(inc, "Rent Expense Supplemental"),
        "ocf": _year_map(cf, "Operating Cash Flow"),
        "capex": _year_map(cf, "Capital Expenditure"),
        "fcf": _year_map(cf, "Free Cash Flow"),
    }


def _ratios_for_year(M, y):
    g = lambda k: M.get(k, {}).get(y)
    revenue = g("revenue"); ac = g("assets_current"); lc = g("liabilities_current")
    cash = g("cash"); sti = g("sti") or 0; at = g("assets_total")
    dcur = g("debt_current") or 0; dnc = g("debt_noncurrent"); lease = g("lease_total") or 0
    equity = g("equity"); oi = g("operating_income")
    ie = g("interest_expense"); ie = abs(ie) if ie is not None else None
    da = g("da"); olc = g("op_lease_cost")
    total_debt = None if dnc is None else (dcur + dnc + lease)
    net_debt = None if total_debt is None else total_debt - (cash or 0) - sti
    ebitda = (oi + da) if (oi is not None and da is not None) else None
    if ebitda is not None and olc is not None:
        ebitdar = ebitda + olc
        lab = "EBITDAR"
    else:
        ebitdar = ebitda
        lab = "EBITDA"
    ratios = {}; nm = {4: False, 6: False}
    ratios[1] = safe_div(ac, lc)
    qn = None if cash is None else (cash + sti)
    ratios[2] = safe_div(qn, lc)
    r3 = safe_div(cash, revenue); ratios[3] = None if r3 is None else r3 * 100
    if ebitdar is None:
        ratios[4] = None
    elif ebitdar <= 0:
        ratios[4] = None; nm[4] = True
    else:
        ratios[4] = safe_div(net_debt, ebitdar)
    ratios[5] = safe_div(total_debt, at)
    if equity is None:
        ratios[6] = None
    elif equity <= 0:
        ratios[6] = None; nm[6] = True
    else:
        ratios[6] = safe_div(total_debt, equity)
    ratios[7] = safe_div(oi, ie)
    ratios[8] = safe_div(ebitdar, ie)
    return {"fy": y, "ratios": ratios, "nm": nm, "ebitda": ebitda, "ebitdar": ebitdar,
            "ebitdar_label": lab, "equity": equity, "revenue": revenue,
            "total_debt": total_debt, "net_debt": net_debt, "cash": cash,
            "assets_total": at, "interest_expense": ie}


def _anchor_years(M):
    ys = set()
    for k in ("assets_total", "revenue", "equity", "operating_income"):
        ys |= set(M.get(k, {}).keys())
    return sorted(ys)


def _fcf_year(M, y):
    f = M.get("fcf", {}).get(y)
    if f is not None:
        return f
    ocf = M.get("ocf", {}).get(y); capex = M.get("capex", {}).get(y)
    if ocf is not None and capex is not None:
        return ocf - abs(capex)
    return None


def build_airline_from_yf(name, ticker, bundle):
    M = extract_yf_financials(bundle)
    ay = _anchor_years(M)
    rec_years = ay[-3:]
    records = [_ratios_for_year(M, y) for y in rec_years]
    label = "EBITDAR" if any(M.get("op_lease_cost", {}).get(y) is not None
                             for y in rec_years) else "EBITDA"
    series_years = ay[-5:]
    series = {
        "years": series_years,
        "current_ratio": {y: safe_div(M.get("assets_current", {}).get(y),
                                       M.get("liabilities_current", {}).get(y))
                          for y in series_years},
        "revenue": {y: M.get("revenue", {}).get(y) for y in series_years},
        "ebitdar": {y: _ratios_for_year(M, y)["ebitdar"] for y in series_years},
        "fcf": {y: _fcf_year(M, y) for y in series_years},
        "ebitdar_label": label,
    }
    pm = None
    hist = bundle.get("history")
    if hist is not None and not hist.empty and "Close" in hist.columns:
        try:
            s = hist["Close"].resample("ME").last().dropna()
            if len(s) > 0 and s.iloc[0] not in (0, None):
                pm = (s / s.iloc[0]) * 100.0
        except Exception:
            pm = None
    info = bundle.get("info") or {}
    stock = {
        "price": info.get("currentPrice") or info.get("regularMarketPrice"),
        "prev": info.get("previousClose"),
        "mcap": info.get("marketCap"),
    }
    return {"ticker": ticker, "label": label, "records": records, "years": rec_years,
            "series": series, "price_monthly": pm, "stock": stock, "bundle": bundle}


def openpyxl_fill(color_name):
    bg, _ = COLOR_MAP[color_name]
    hexv = bg.replace("#", "")
    return PatternFill(start_color=hexv, end_color=hexv, fill_type="solid")


def build_excel():
    wb = Workbook()
    bold = Font(bold=True)

    # Sheet 1: Credit Summary (latest FY) with RAG fills
    ws = wb.active
    ws.title = "Credit Summary"
    ws.cell(row=1, column=1, value="Ratio").font = bold
    for j, a in enumerate(airlines, start=2):
        ws.cell(row=1, column=j, value=a).font = bold
    r = 2
    last_dim = None
    for rid, label, key, sym, dim in RATIOS:
        if dim != last_dim:
            ws.cell(row=r, column=1, value=dim).font = bold
            r += 1
            last_dim = dim
        ws.cell(row=r, column=1, value=label)
        for j, a in enumerate(airlines, start=2):
            recs = data[a]["records"]
            v = recs[-1]["ratios"].get(rid) if recs else None
            cell = ws.cell(row=r, column=j, value=fmt_ratio(rid, v))
            cname = resolve_color(rid, recs, peers_latest[rid], peer_mode)
            cell.fill = openpyxl_fill(cname)
            cell.alignment = Alignment(horizontal="center")
        r += 1
    ws.column_dimensions["A"].width = 24
    for j in range(2, 2 + len(airlines)):
        ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = 18

    # Sheet 2: 3-Year Trends
    ws2 = wb.create_sheet("3-Year Trends")
    headers = ["Airline", "Ticker", "FY"] + [r[1] for r in RATIOS]
    for j, h in enumerate(headers, start=1):
        ws2.cell(row=1, column=j, value=h).font = bold
    row = 2
    for a in airlines:
        for rec in data[a]["records"]:
            ws2.cell(row=row, column=1, value=a)
            ws2.cell(row=row, column=2, value=data[a]["ticker"])
            ws2.cell(row=row, column=3, value=f"FY{rec['fy']}")
            for k, (rid, *_rest) in enumerate(RATIOS, start=4):
                ws2.cell(row=row, column=k, value=fmt_ratio(rid, rec["ratios"].get(rid)))
            row += 1

    # Sheet 3: Raw Financials ($M)
    ws3 = wb.create_sheet("Raw Financials")
    rawhead = ["Airline", "Ticker", "FY", "Revenue ($M)", "EBITDA(R)", "EBITDA(R) ($M)",
               "Total Debt ($M)", "Net Debt ($M)", "Cash ($M)", "Total Assets ($M)",
               "Equity ($M)", "Interest Expense ($M)"]
    for j, h in enumerate(rawhead, start=1):
        ws3.cell(row=1, column=j, value=h).font = bold
    row = 2
    for a in airlines:
        for rec in data[a]["records"]:
            vals = [
                a, data[a]["ticker"], f"FY{rec['fy']}",
                to_millions(rec["revenue"]), rec["ebitdar_label"], to_millions(rec["ebitdar"]),
                to_millions(rec["total_debt"]), to_millions(rec["net_debt"]),
                to_millions(rec["cash"]), to_millions(rec["assets_total"]),
                to_millions(rec["equity"]), to_millions(rec["interest_expense"]),
            ]
            for j, v in enumerate(vals, start=1):
                if isinstance(v, float):
                    ws3.cell(row=row, column=j, value=round(v, 1))
                else:
                    ws3.cell(row=row, column=j, value=v)
            row += 1

    # Sheet 4: Methodology
    ws4 = wb.create_sheet("Methodology")
    ws4.column_dimensions["A"].width = 110
    lines = [
        "AVIATION FINANCE DASHBOARD \u2014 METHODOLOGY",
        "",
        "DATA SOURCE",
        "All figures sourced from Yahoo Finance (yfinance) statement data.",
        "This tool now supports global publicly listed airlines, not limited to U.S. SEC filers.",
        "Cross-border peer comparisons may mix IFRS and GAAP reporting standards — ratios should",
        "be interpreted with this in mind when comparing airlines across accounting regimes.",
        "",
        "DERIVED METRICS",
        "Total Debt = current debt + long-term debt + operating lease liability (combined, per yfinance)",
        "Net Debt = Total Debt - Cash - Short-term Investments",
        "EBITDA = Operating Income + Depreciation & Amortisation",
        "EBITDAR = EBITDA + Operating Lease Cost (when reported; otherwise EBITDA is used).",
        "Note: yfinance rarely exposes lease cost/liability, so most airlines fall back to EBITDA.",
        "",
        "RATIOS, LESSOR PURPOSE, AND THRESHOLD RATIONALE",
        "1. Current Ratio = Current Assets / Current Liabilities",
        "   Purpose: near-term ability to meet obligations incl. lease rentals. Green >=0.90, Amber 0.60-0.90, Red <0.60.",
        "2. Quick Ratio = (Cash + Short-term Investments) / Current Liabilities",
        "   Purpose: liquidity excluding less-liquid current assets. Green >=0.80, Amber 0.50-0.80, Red <0.50.",
        "3. Cash % Revenue = Cash / Revenue x 100",
        "   Purpose: liquidity buffer relative to operating scale. Green >=15, Amber 10-15, Red <10.",
        "4. Net Debt / EBITDAR = Net Debt / EBITDAR",
        "   Purpose: lease-adjusted leverage \u2014 the core lessor metric. Green <=4.0, Amber 4.0-6.0, Red >6.0.",
        "5. Total Debt / Assets = Total Debt / Total Assets",
        "   Purpose: balance-sheet leverage. Green <=0.70, Amber 0.70-0.85, Red >0.85.",
        "6. Total Debt / Equity = Total Debt / Stockholders' Equity",
        "   Purpose: leverage vs equity cushion. Green <=3.0, Amber 3.0-5.0, Red >5.0.",
        "7. Interest Coverage = Operating Income / Interest Expense",
        "   Purpose: ability to service interest. Green >=2.0, Amber 1.0-2.0, Red <1.0.",
        "8. EBITDAR Coverage = EBITDAR / Interest Expense",
        "   Purpose: lease-adjusted earnings vs interest. Green >=3.0, Amber 2.0-3.0, Red <2.0.",
        "",
        "EBITDAR vs EBITDA",
        "EBITDAR adds back operating lease cost to allow comparison across airlines with different",
        "owned/leased fleet mixes. When operating lease cost is not separately reported, EBITDA is used",
        "and the metric is labelled accordingly per airline.",
        "",
        "N/M TREATMENT",
        "Net Debt / EBITDAR is Not Meaningful when EBITDAR <= 0 (a leverage multiple over negative",
        "earnings is uninformative). Total Debt / Equity is Not Meaningful when equity <= 0 (deficit).",
        "For the latest year, N/M cells are coloured by trend: deteriorating/persistent N/M = Red;",
        "improving toward positive = Amber. N/M only in prior years is coloured normally on the valid latest year.",
        "",
        "PEER COMPARISON RULE",
        "With 4+ airlines selected, RAG bands use peer median/quartile logic (leverage: below median = green,",
        "top quartile = red; others: above median = green, bottom quartile = red). With fewer than 4 airlines,",
        "fixed absolute thresholds are used.",
        "",
        "Verify all figures against primary filings before any investment or credit decision.",
    ]
    for i, ln in enumerate(lines, start=1):
        c = ws4.cell(row=i, column=1, value=ln)
        if ln and ln.isupper() and len(ln) < 60:
            c.font = bold

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ----------------------------------------------------------------------------
# Sidebar: search + add tickers
# ----------------------------------------------------------------------------
st.sidebar.title("\u2708 Airline Selection")

if "data" not in st.session_state:
    st.session_state.data = {}
if "add_msg" not in st.session_state:
    st.session_state.add_msg = None


def _add_ticker(ticker, prefer_name=None):
    ticker = (ticker or "").strip()
    if not ticker:
        return
    up = ticker.upper()
    for nm, d in st.session_state.data.items():
        if d.get("ticker") == up:
            st.session_state.add_msg = ("info", f"{nm} ({up}) is already loaded.")
            return
    with st.spinner(f"Fetching {ticker} ..."):
        bundle = load_ticker_bundle(ticker)
    info = bundle.get("info") or {}
    name = info.get("longName") or info.get("shortName") or prefer_name
    if not name:
        st.session_state.add_msg = (
            "error", f"Could not find data for '{ticker}'. Check the ticker symbol and try again.")
        return
    st.session_state.data[name] = build_airline_from_yf(name, up, bundle)
    st.session_state.add_msg = ("success", f"Added {name} ({up}).")


query = st.sidebar.text_input("Search company name or ticker", key="search_box")
if query:
    q = query.strip().lower()
    matches = [(n, t) for n, t in AIRLINE_SUGGESTIONS.items()
               if q in n.lower() or q in t.lower()][:8]
    for n, t in matches:
        if st.sidebar.button(f"{n}  ({t})", key=f"sug_{t}", use_container_width=True):
            _add_ticker(t, prefer_name=n)
            st.rerun()

if st.sidebar.button("Add Ticker", key="add_btn", use_container_width=True):
    _add_ticker(query)
    st.rerun()

if st.session_state.add_msg:
    kind, msg = st.session_state.add_msg
    getattr(st.sidebar, kind)(msg)
    st.session_state.add_msg = None

data = st.session_state.data
peer_mode = len(data) >= 4

if data:
    mode_label = "peer comparison" if peer_mode else "absolute thresholds"
    st.sidebar.success(f"{len(data)} airline(s) loaded \u2014 using {mode_label}.")
    if not peer_mode:
        st.sidebar.warning("Peer comparison requires 4+ airlines. Using absolute thresholds.")
    st.sidebar.markdown("---")
    remove = None
    for name in list(data.keys()):
        stock = data[name].get("stock") or {}
        ticker = data[name]["ticker"]
        col1, col2 = st.sidebar.columns([6, 1])
        col1.markdown(
            f"<div style='line-height:1.3;'>"
            f"<div style='font-weight:600;font-size:13px;'>{name} "
            f"<span style='color:#6b7280;'>({ticker})</span></div>"
            f"<div style='font-size:11px;'>{price_line_html(stock)}</div></div>",
            unsafe_allow_html=True,
        )
        if col2.button("\u00d7", key=f"rm_{ticker}", help=f"Remove {name}"):
            remove = name
    if remove is not None:
        del st.session_state.data[remove]
        st.rerun()
    st.sidebar.markdown(
        "<div style='font-size:10px;color:#9ca3af;font-style:italic;margin-top:6px;'>"
        "Prices delayed ~15 min. Source: Yahoo Finance. Financial statement data sourced "
        "from Yahoo Finance. Reporting standards vary by country (GAAP/IFRS).</div>",
        unsafe_allow_html=True,
    )


# ----------------------------------------------------------------------------
# Header
# ----------------------------------------------------------------------------
st.title("Aviation Finance Dashboard")
st.caption("Global Airline Credit & Financials | Data: Yahoo Finance (yfinance)")

if not data:
    st.info("Search for an airline or type a ticker in the sidebar, then pick a suggestion "
            "or click **Add Ticker** to begin.")
    st.stop()

airlines = list(data.keys())
color_seq = px.colors.qualitative.Plotly
airline_colors = {a: color_seq[i % len(color_seq)] for i, a in enumerate(airlines)}

peers_latest = {
    rid: [data[a]["records"][-1]["ratios"].get(rid) if data[a]["records"] else None
          for a in airlines]
    for rid, *_ in RATIOS
}


def latest_fy_label():
    yrs = [d["years"][-1] for d in data.values() if d.get("years")]
    return f"FY{max(yrs)}" if yrs else "Latest FY"


export_bytes = build_excel()
st.download_button(
    "\u2b07 Export to Excel",
    data=export_bytes,
    file_name=f"Aviation_Finance_Analysis_{date.today().isoformat()}.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)


# ----------------------------------------------------------------------------
# Fleet Composition engine (live Wikipedia scrape)
# ----------------------------------------------------------------------------
WIKI_HEADERS = {"User-Agent": "Mozilla/5.0 (Aviation Finance Dashboard)"}
TICKER_TO_SUGGESTION = {v: k for k, v in AIRLINE_SUGGESTIONS.items()}
AC_MAKERS = ("boeing", "airbus", "atr", "embraer", "bombardier", "mcdonnell douglas",
             "de havilland", "comac", "sukhoi", "antonov", "saab", "fokker")
# Header substrings that mark an infobox / non-fleet table and must cause rejection.
FLEET_REJECT = ("isin", "employees", "destinations", "revenue", "traded as",
                "subsidiaries", "commenced operations", "headquarters", "founded",
                "operating income", "net income", "total assets", "website",
                "frequent-flyer", "alliance", "parent company")


def _clean_text(x):
    s = re.sub(r"\[[^\]]*\]", "", str(x))      # footnote markers [1], [a]
    return re.sub(r"\s+", " ", s).strip()


def _to_count(x):
    s = re.sub(r"\[[^\]]*\]", "", str(x)).replace(",", "")
    m = re.search(r"-?\d+", s)
    return (int(m.group()), False) if m else (0, True)


def _norm_cols(df):
    cols = []
    for c in df.columns:
        if isinstance(c, tuple):
            c = " ".join(str(p) for p in c if str(p) != "nan")
        cols.append(_clean_text(c).lower())
    return cols


def _has_maker(text):
    t = text.lower()
    return any(m in t for m in AC_MAKERS)


def _strict_fleet_cols(df):
    """Strict identification. Returns (aircraft_idx, in_service_idx) or None.

    Requires an 'Aircraft' header AND an 'In service'/'In fleet' header, rejects
    any table carrying infobox-style headers, and never selects an 'Orders' column.
    """
    if df.shape[1] < 2:
        return None
    cols = _norm_cols(df)
    if any(any(bad in h for bad in FLEET_REJECT) for h in cols):
        return None
    ac_idx = next((i for i, h in enumerate(cols) if "aircraft" in h), None)
    if ac_idx is None:
        return None
    cnt_idx = next((i for i, h in enumerate(cols)
                    if i != ac_idx and "order" not in h
                    and ("in service" in h or "in fleet" in h)), None)
    if cnt_idx is None:
        return None
    return ac_idx, cnt_idx


def parse_fleet(tables):
    """Return (DataFrame|None, total, flagged) for the first table passing strict checks."""
    for df in tables:
        ident = _strict_fleet_cols(df)
        if not ident:
            continue
        ac_idx, cnt_idx = ident
        rows, flagged = [], False
        for _, r in df.iterrows():
            ac = _clean_text(r.iloc[ac_idx])
            if not ac or ac.lower() == "nan" or "total" in ac.lower():
                continue
            if not _has_maker(ac):           # drop summary / stray rows
                continue
            cnt, flag = _to_count(r.iloc[cnt_idx])
            flagged = flagged or flag
            rows.append((ac, cnt))
        if len(rows) < 3:                    # a real fleet table lists several types
            continue
        out = pd.DataFrame(rows, columns=["Aircraft Type", "In Service"])
        # merged multi-row variants repeat a model: keep one row per type
        out = out.drop_duplicates(subset="Aircraft Type", keep="first").reset_index(drop=True)
        if len(out) < 3:
            continue
        out = out.sort_values("In Service", ascending=False).reset_index(drop=True)
        return out, int(out["In Service"].sum()), flagged
    return None, 0, False


def _table_after_fleet_heading(html):
    """Return the HTML of the first <table> following the page's 'Fleet' heading."""
    try:
        import lxml.html as LH
        doc = LH.fromstring(html)
    except Exception:
        return None
    els = list(doc.iter())
    start = None
    for i, el in enumerate(els):
        tag = el.tag if isinstance(el.tag, str) else ""
        if (el.get("id") or "").strip().lower() == "fleet":
            start = i
            break
        if tag in ("h2", "h3"):
            txt = (el.text_content() or "").strip().lower()
            if txt == "fleet" or txt.startswith("fleet "):
                start = i
                break
    if start is None:
        return None
    for el in els[start + 1:]:
        tag = el.tag if isinstance(el.tag, str) else ""
        if tag == "h2":                      # next top-level section: stop before it
            break
        if tag == "table":
            try:
                return LH.tostring(el, encoding="unicode")
            except Exception:
                return None
    return None


def _wiki_url(name):
    return "https://en.wikipedia.org/wiki/" + name.replace(" ", "_")


def fleet_name_candidates(display_name, ticker):
    cands = []
    sug = TICKER_TO_SUGGESTION.get(ticker)
    for nm in (sug, display_name):
        if nm:
            cands.append(re.sub(r"\s*\(.*?\)", "", nm).strip())
    if display_name:
        cleaned = re.sub(
            r",?\s*(Inc\.?|Incorporated|Corporation|Corp\.?|plc|PLC|Holdings|Company|"
            r"Co\.?|Group|Ltd\.?|Limited|S\.A\.?|AG|N\.V\.?)\b.*$", "", display_name).strip()
        if cleaned:
            cands.append(cleaned)
    seen, out = set(), []
    for c in cands:
        if c and c.lower() not in seen:
            seen.add(c.lower())
            out.append(c)
    return out


def _resolve_wiki_title(query):
    try:
        r = requests.get(
            "https://en.wikipedia.org/w/api.php",
            params={"action": "opensearch", "search": query, "limit": 1,
                    "namespace": 0, "format": "json"},
            headers=WIKI_HEADERS, timeout=10)
        data = r.json()
        if isinstance(data, list) and len(data) >= 4 and data[3]:
            return data[3][0]
        if isinstance(data, list) and len(data) >= 2 and data[1]:
            return _wiki_url(data[1][0])
    except Exception:
        return None
    return None


def _read_tables(html):
    try:
        return pd.read_html(io.StringIO(html))
    except Exception:
        return []


def _get(url):
    resp = requests.get(url, headers=WIKI_HEADERS, timeout=10)
    return resp.text if resp.status_code == 200 else None


def scrape_fleet(candidates):
    """Dedicated _fleet page -> main-page Fleet-section table -> opensearch fallback."""
    primary = (_wiki_url(candidates[0]) + "_fleet") if candidates else "https://en.wikipedia.org/"

    # 1) dedicated "<Airline>_fleet" pages (scan tables, strict validator rejects non-fleet)
    for nm in candidates:
        url = _wiki_url(nm) + "_fleet"
        try:
            html = _get(url)
            if html:
                df, total, flagged = parse_fleet(_read_tables(html))
                if df is not None:
                    return {"ok": True, "df": df, "total": total, "url": url, "flagged": flagged}
        except Exception:
            pass

    # 2) main page: ONLY the table that follows the "Fleet" heading (no indiscriminate scan)
    for nm in candidates:
        url = _wiki_url(nm)
        try:
            html = _get(url)
            if html:
                frag = _table_after_fleet_heading(html)
                if frag:
                    df, total, flagged = parse_fleet(_read_tables(frag))
                    if df is not None:
                        return {"ok": True, "df": df, "total": total, "url": url, "flagged": flagged}
        except Exception:
            pass

    # 3) opensearch fallback: resolve a real article for "<name> fleet" then "<name>"
    if candidates:
        for q in (candidates[0] + " fleet", candidates[0]):
            resolved = _resolve_wiki_title(q)
            if not resolved:
                continue
            try:
                html = _get(resolved)
                if not html:
                    continue
                frag = _table_after_fleet_heading(html)
                tables = _read_tables(frag) if frag else _read_tables(html)
                df, total, flagged = parse_fleet(tables)
                if df is not None:
                    return {"ok": True, "df": df, "total": total, "url": resolved, "flagged": flagged}
            except Exception:
                pass

    return {"ok": False, "url": primary}


# ----------------------------------------------------------------------------
# Reverse DCF helpers
# ----------------------------------------------------------------------------
@st.cache_data(show_spinner=False, ttl=300)
def fetch_risk_free():
    """10Y Treasury yield (^TNX quotes in percent) as a decimal, or None."""
    try:
        h = yf.Ticker("^TNX").history(period="1d")
        return float(h["Close"].iloc[-1]) / 100.0
    except Exception:
        return None


def _effective_tax_rate(bundle):
    """Avg of yearly (tax / pretax) over up to 3 FYs, excluding loss years.
    Returns (rate, used_fallback). Clipped to [0, 0.35]; 0.21 fallback otherwise."""
    inc = bundle.get("fin")
    tax = _year_map(inc, "Tax Provision")
    pre = _year_map(inc, "Pretax Income")
    years = sorted(set(tax) & set(pre), reverse=True)[:3]
    rates = [tax[y] / pre[y] for y in years
             if tax.get(y) is not None and pre.get(y) is not None and pre[y] > 0]
    if not rates:
        return 0.21, True
    avg = sum(rates) / len(rates)
    if avg < 0.0 or avg > 0.35:
        return 0.21, True
    return avg, False


def _ttm_fcf(bundle):
    """TTM unlevered-FCF proxy = sum(OCF) + sum(CapEx) over up to 4 latest quarters.
    Returns (ttm_value_or_None, n_quarters_used)."""
    qcf = bundle.get("qcf")
    if qcf is None or qcf.empty:
        return None, 0
    ocf_idx = _yf_find(qcf, "Operating Cash Flow")
    cap_idx = _yf_find(qcf, "Capital Expenditure")
    if ocf_idx is None:
        return None, 0
    ocf_sum, cap_sum, n = 0.0, 0.0, 0
    for c in list(qcf.columns)[:4]:
        try:
            v = qcf.loc[ocf_idx, c]
        except Exception:
            continue
        if pd.isna(v):
            continue
        ocf_sum += float(v)
        n += 1
        if cap_idx is not None:
            try:
                cv = qcf.loc[cap_idx, c]
                if not pd.isna(cv):
                    cap_sum += float(cv)
            except Exception:
                pass
    if n == 0:
        return None, 0
    ttm = ocf_sum + (cap_sum if cap_sum <= 0 else -cap_sum)
    return ttm, n


def _money(v):
    if v is None:
        return "\u2014"
    a = abs(v)
    sign = "-" if v < 0 else ""
    if a >= 1e9:
        return f"{sign}${a / 1e9:.1f}B"
    if a >= 1e6:
        return f"{sign}${a / 1e6:.1f}M"
    return f"{sign}${a:,.0f}"


def _pct(x):
    return "\u2014" if x is None else f"{x * 100:.2f}%"


# ----------------------------------------------------------------------------
# Fuel Analysis engine (EIA jet fuel + EBITDAR correlation + EDGAR sensitivity)
# ----------------------------------------------------------------------------
# NOTE: The EIA API key is read from Streamlit secrets (st.secrets["EIA_API_KEY"]).
# Set it in .streamlit/secrets.toml (or the app's Secrets settings) as:
#     EIA_API_KEY = "your_key"
# Register a free key at https://api.eia.gov/opendata/register. If the secret is
# missing, Section 1/2 degrade to the "could not load" path (the app still runs).
EIA_JET_FUEL_URL_TEMPLATE = (
    "https://api.eia.gov/v2/petroleum/pri/wfr/data/"
    "?api_key={key}&frequency=weekly&data[0]=value"
    "&facets[product][]=EPJ&facets[duoarea][]=R30"
    "&sort[0][column]=period&sort[0][direction]=desc&length=200"
)
US_EXCHANGES = {"NMS", "NYQ", "NGM", "NCM", "ASE"}
US_CARRIER_CIKS = {
    "DAL": "0000027904", "UAL": "0000100517", "AAL": "0001549922",
    "LUV": "0000092380", "ALK": "0000766421", "JBLU": "0001158463",
    "ALGT": "0001362988", "ULCC": "0001670076", "SAVE": "0001418121",
    "SNCY": "0001549802", "HA": "0000046619", "SKYW": "0000070858",
    "ATSG": "0000894871", "MESA": "0000810332",
}
EDGAR_HEADERS = {"User-Agent": "Aviation Finance Dashboard contact@aviationdashboard.com"}


@st.cache_data(show_spinner=False, ttl=3600)
def fetch_eia_jet_fuel():
    try:
        eia_key = st.secrets["EIA_API_KEY"]
    except Exception:
        return None
    try:
        st.write(f"DEBUG EIA URL: {EIA_JET_FUEL_URL_TEMPLATE.format(key=eia_key)[:100]}...")
        resp = requests.get(EIA_JET_FUEL_URL_TEMPLATE.format(key=eia_key), timeout=10)
        rows = resp.json()["response"]["data"]
        df = pd.DataFrame(rows)
        df["period"] = pd.to_datetime(df["period"])
        df["value"] = pd.to_numeric(df["value"], errors="coerce")
        df = df.dropna(subset=["value"]).sort_values("period").reset_index(drop=True)
        return df if not df.empty else None
    except Exception as e:
        import traceback
        st.write(f"DEBUG EIA ERROR: {type(e).__name__}: {e}")
        st.write(traceback.format_exc())
        return None


def _filter_fuel_timeframe(df, choice):
    if df is None or df.empty:
        return df
    end = df["period"].max()
    today = pd.Timestamp.today().normalize()
    if choice == "Year to Date":
        start = pd.Timestamp(year=today.year, month=1, day=1)
    elif choice == "Last Month":
        start = end - pd.Timedelta(days=30)
    elif choice == "Last 12 Months":
        start = end - pd.Timedelta(days=365)
    else:
        start = end - pd.Timedelta(days=1095)
    return df[df["period"] >= start]


def _quarter_ebitdar_margins(bundle):
    qi = bundle.get("qfin")
    qc = bundle.get("qcf")
    if qi is None or getattr(qi, "empty", True):
        return pd.DataFrame()
    oi_idx = _yf_find(qi, "Operating Income")
    rev_idx = _yf_find(qi, "Total Revenue")
    if oi_idx is None or rev_idx is None:
        return pd.DataFrame()
    da_idx = _yf_find(qc, "Depreciation And Amortization") if qc is not None else None
    lease_idx = _yf_find(qi, "Rent Expense Supplemental")
    rows = []
    for c in list(qi.columns)[:4]:
        try:
            oi = qi.loc[oi_idx, c]
            rev = qi.loc[rev_idx, c]
        except Exception:
            continue
        if pd.isna(oi) or pd.isna(rev) or rev == 0:
            continue
        da = 0.0
        if da_idx is not None and qc is not None and c in qc.columns:
            dv = qc.loc[da_idx, c]
            da = 0.0 if pd.isna(dv) else float(dv)
        lease = 0.0
        if lease_idx is not None:
            lv = qi.loc[lease_idx, c]
            lease = 0.0 if pd.isna(lv) else float(lv)
        ebitdar = float(oi) + da + lease
        rows.append({"q_end": pd.Timestamp(c),
                     "label": _col_label(pd.Timestamp(c), "Quarterly"),
                     "margin": ebitdar / float(rev) * 100})
    return pd.DataFrame(rows)


def _avg_fuel_for_quarter(df_fuel, q_end):
    start = q_end - pd.Timedelta(days=90)
    sub = df_fuel[(df_fuel["period"] >= start) & (df_fuel["period"] <= q_end)]
    return sub["value"].mean() if not sub.empty else None


def _corr(xs, ys):
    try:
        from scipy import stats
        res = stats.linregress(xs, ys)
        return float(res.rvalue), float(res.slope), float(res.intercept)
    except Exception:
        import numpy as np
        if len(xs) < 2:
            return None, None, None
        r = float(np.corrcoef(xs, ys)[0, 1])
        slope, intercept = (float(v) for v in np.polyfit(xs, ys, 1))
        return r, slope, intercept


@st.cache_data(show_spinner=False, ttl=3600)
def fetch_edgar_fuel(cik):
    """(fuel_cost, opex) for the most recent FY from SEC EDGAR, or (None, None)."""
    try:
        url = f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
        resp = requests.get(url, headers=EDGAR_HEADERS, timeout=10)
        facts = resp.json().get("facts", {}).get("us-gaap", {})
    except Exception:
        return None, None

    def latest_annual(concepts):
        for c in concepts:
            node = facts.get(c)
            if not node:
                continue
            for _unit, series in node.get("units", {}).items():
                best = None
                for e in series:
                    if e.get("form") != "10-K" or e.get("fp") != "FY" or e.get("val") is None:
                        continue
                    if best is None or e.get("end", "") > best.get("end", ""):
                        best = e
                if best:
                    return float(best["val"])
        return None

    fuel = latest_annual(["AirlineFuelCosts",
                          "FuelCosts",
                          "FuelCostsAndTaxes",
                          "AircraftFuelAndRelatedTaxes",
                          "AirlineCapacityPurchaseArrangements"])
    opex = latest_annual(["OperatingExpenses",
                          "CostsAndExpenses",
                          "OperatingCostsAndExpenses",
                          "CostsAndExpensesTotal"])
    return fuel, opex


def _is_us_carrier(ticker, info):
    exch = info.get("exchange") or info.get("fullExchangeName") or ""
    return exch in US_EXCHANGES or "." not in ticker


tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "Credit Summary", "Trends", "Peer Ranking", "Financial Statements",
    "Fleet Composition", "Reverse DCF", "Fuel Analysis"
])


# ----------------------------------------------------------------------------
# TAB 1 \u2014 Credit Summary
# ----------------------------------------------------------------------------
with tab1:
    st.subheader(f"Credit Summary \u2014 {latest_fy_label()}")

    # Signature element: RAG proportion bar per airline
    st.markdown("**Metric health (latest FY)**")
    for a in airlines:
        recs = data[a]["records"]
        counts = {"green": 0, "amber": 0, "red": 0, "nm": 0}
        for rid, *_ in RATIOS:
            counts[resolve_color(rid, recs, peers_latest[rid], peer_mode)] += 1
        total = sum(counts.values()) or 1
        seg = ""
        for cname in ["green", "amber", "red", "nm"]:
            if counts[cname] == 0:
                continue
            w = counts[cname] / total * 100
            bg, txt = COLOR_MAP[cname]
            border = txt
            seg += (
                f"<div style='width:{w:.1f}%;background:{bg};color:{txt};"
                f"border-right:1px solid #fff;text-align:center;font-size:12px;"
                f"line-height:22px;border:1px solid {border}33;'>{counts[cname]}</div>"
            )
        bar = (
            f"<div style='display:flex;width:100%;height:22px;border-radius:4px;"
            f"overflow:hidden;margin-bottom:6px;'>{seg}</div>"
        )
        st.markdown(
            f"<div style='font-size:13px;margin-bottom:2px;'>{a} "
            f"<span style='color:#6b7280;'>({data[a]['ticker']})</span></div>{bar}",
            unsafe_allow_html=True,
        )

    st.markdown("---")

    cols = list(airlines)
    show_median = peer_mode
    if show_median:
        cols = cols + ["Peer Median"]

    index_rows = []
    text_grid = {}
    color_grid = {}
    last_dim = None
    for rid, label, key, sym, dim in RATIOS:
        if dim != last_dim:
            index_rows.append(dim)
            text_grid[dim] = {c: "" for c in cols}
            color_grid[dim] = {c: f"font-weight:700;background:#eef2ff;color:#1e3a8a" for c in cols}
            last_dim = dim
        index_rows.append(label)
        text_grid[label] = {}
        color_grid[label] = {}
        for a in airlines:
            recs = data[a]["records"]
            v = recs[-1]["ratios"].get(rid) if recs else None
            text_grid[label][a] = fmt_ratio(rid, v)
            cname = resolve_color(rid, recs, peers_latest[rid], peer_mode)
            bg, txt = COLOR_MAP[cname]
            color_grid[label][a] = f"background:{bg};color:{txt};text-align:center"
        if show_median:
            vals = [v for v in peers_latest[rid] if v is not None]
            med = statistics.median(vals) if vals else None
            text_grid[label]["Peer Median"] = fmt_ratio(rid, med)
            color_grid[label]["Peer Median"] = "background:#f1f5f9;color:#334155;text-align:center;font-style:italic"

    df_disp = pd.DataFrame(
        [[text_grid[r].get(c, "") for c in cols] for r in index_rows],
        index=index_rows, columns=cols,
    )
    css_df = pd.DataFrame(
        [[color_grid[r].get(c, "") for c in cols] for r in index_rows],
        index=index_rows, columns=cols,
    )
    styler = df_disp.style.apply(lambda _: css_df, axis=None)
    st.dataframe(styler, use_container_width=True)
    st.caption("Ratios shown as Nx; Cash % Revenue as %. N/M = not meaningful. "
               "Colour bands: " + ("peer median/quartile." if peer_mode else "absolute thresholds."))

# ----------------------------------------------------------------------------
# TAB 2 \u2014 Trends (5 charts, single column, up to 5y history)
# ----------------------------------------------------------------------------
with tab2:
    st.subheader("Trends")

    for a in airlines:
        ys = data[a]["series"]["years"]
        if ys and len(ys) < 5:
            st.caption(f"Limited history available for {a} \u2014 showing {len(ys)} year(s).")

    def _fy_order(dfp):
        return sorted(dfp["FY"].unique())

    # 1) Current Ratio
    st.markdown("**Current Ratio** \u2014 liquidity vs short-term obligations "
                "(green dashed line = 0.90 threshold).")
    rows = []
    for a in airlines:
        s = data[a]["series"]
        for y in s["years"]:
            rows.append({"Airline": a, "FY": f"FY{y}", "Value": s["current_ratio"].get(y)})
    dfp = pd.DataFrame(rows)
    if not dfp.empty and not dfp["Value"].dropna().empty:
        fig = px.line(dfp, x="FY", y="Value", color="Airline", markers=True,
                      category_orders={"FY": _fy_order(dfp)}, color_discrete_map=airline_colors)
        fig.add_hline(y=0.90, line_dash="dash", line_color="#15803d",
                      annotation_text="0.90", annotation_position="top left")
        fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10), yaxis_title="x")
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("No current-ratio data available.")

    # 2) Revenue Growth %
    st.markdown("**Revenue Growth %** \u2014 year-over-year change in total revenue "
                "(0% reference line).")
    rows = []
    for a in airlines:
        s = data[a]["series"]; ys = s["years"]
        for i, y in enumerate(ys):
            if i == 0:
                continue
            cur = s["revenue"].get(y); prev = s["revenue"].get(ys[i - 1])
            val = ((cur - prev) / prev * 100) if (cur is not None and prev not in (None, 0)) else None
            rows.append({"Airline": a, "FY": f"FY{y}", "Value": val})
    dfp = pd.DataFrame(rows)
    if not dfp.empty and not dfp["Value"].dropna().empty:
        fig = px.line(dfp, x="FY", y="Value", color="Airline", markers=True,
                      category_orders={"FY": _fy_order(dfp)}, color_discrete_map=airline_colors)
        fig.add_hline(y=0, line_dash="dash", line_color="#6b7280")
        fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10), yaxis_title="% YoY")
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Not enough history for revenue growth.")

    # 3) EBITDA(R) Growth %
    st.markdown("**EBITDA(R) Growth %** \u2014 year-over-year change in lease-adjusted earnings "
                "(0% reference line).")
    rows = []
    for a in airlines:
        s = data[a]["series"]; ys = s["years"]
        for i, y in enumerate(ys):
            if i == 0:
                continue
            cur = s["ebitdar"].get(y); prev = s["ebitdar"].get(ys[i - 1])
            val = ((cur - prev) / abs(prev) * 100) if (cur is not None and prev not in (None, 0)) else None
            rows.append({"Airline": a, "FY": f"FY{y}", "Value": val})
    dfp = pd.DataFrame(rows)
    if not dfp.empty and not dfp["Value"].dropna().empty:
        fig = px.line(dfp, x="FY", y="Value", color="Airline", markers=True,
                      category_orders={"FY": _fy_order(dfp)}, color_discrete_map=airline_colors)
        fig.add_hline(y=0, line_dash="dash", line_color="#6b7280")
        fig.update_layout(height=360, margin=dict(l=10, r=10, t=10, b=10), yaxis_title="% YoY")
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("Not enough history for EBITDA(R) growth.")

    # 4) Free Cash Flow (bars, sign-coloured)
    st.markdown("**Free Cash Flow** \u2014 absolute FCF per year. Bars are in each airline's "
                "**local reporting currency** (not FX-converted); cross-airline bar heights "
                "are therefore not directly comparable.")
    if len(airlines) >= 4:
        cols = st.columns(2)
        for idx, a in enumerate(airlines):
            s = data[a]["series"]
            xs = [f"FY{y}" for y in s["years"]]
            vm = [(s["fcf"].get(y) / 1e6 if s["fcf"].get(y) is not None else None) for y in s["years"]]
            colors = ["#15803d" if (v is not None and v >= 0) else "#b91c1c" for v in vm]
            fig = go.Figure(go.Bar(x=xs, y=vm, marker_color=colors))
            fig.add_hline(y=0, line_color="#6b7280")
            fig.update_layout(title=a, height=260, margin=dict(l=10, r=10, t=30, b=10),
                              yaxis_title="millions (local)")
            cols[idx % 2].plotly_chart(fig, use_container_width=True)
    else:
        fig = go.Figure()
        for a in airlines:
            s = data[a]["series"]
            xs = [f"FY{y}" for y in s["years"]]
            vm = [(s["fcf"].get(y) / 1e6 if s["fcf"].get(y) is not None else None) for y in s["years"]]
            colors = ["#15803d" if (v is not None and v >= 0) else "#b91c1c" for v in vm]
            fig.add_trace(go.Bar(name=a, x=xs, y=vm, marker_color=colors))
        fig.add_hline(y=0, line_color="#6b7280")
        fig.update_layout(barmode="group", height=380, margin=dict(l=10, r=10, t=10, b=10),
                          yaxis_title="millions (local)")
        st.plotly_chart(fig, use_container_width=True)

    # 5) Indexed stock price
    st.markdown("**Relative Stock Price (indexed to 100)** \u2014 relative performance across airlines.")
    rows = []
    for a in airlines:
        pm = data[a].get("price_monthly")
        if pm is None or len(pm) == 0:
            continue
        for dt, v in pm.items():
            rows.append({"Airline": a, "Date": dt, "Indexed": float(v)})
    dfp = pd.DataFrame(rows)
    if not dfp.empty:
        fig = px.line(dfp, x="Date", y="Indexed", color="Airline", color_discrete_map=airline_colors)
        fig.add_hline(y=100, line_dash="dash", line_color="#6b7280", annotation_text="100")
        fig.update_layout(height=380, margin=dict(l=10, r=10, t=10, b=10),
                          yaxis_title="Indexed (start = 100)")
        st.plotly_chart(fig, use_container_width=True)
        st.caption("Indexed to 100 at start of period. Reflects relative price performance, "
                   "not absolute returns.")
    else:
        st.info("No price history available.")


# ----------------------------------------------------------------------------
# TAB 3 \u2014 Peer Ranking
# ----------------------------------------------------------------------------
with tab3:
    st.subheader(f"Peer Ranking \u2014 {latest_fy_label()}")
    if len(airlines) < 2:
        st.info("Peer ranking needs at least 2 airlines. Load more to compare.")
    else:
        grid = st.columns(2)
        for i, (rid, label, key, sym, dim) in enumerate(RATIOS):
            rows = []
            for a in airlines:
                recs = data[a]["records"]
                v = recs[-1]["ratios"].get(rid) if recs else None
                if v is None:
                    continue
                cname = resolve_color(rid, recs, peers_latest[rid], peer_mode)
                rows.append({"Airline": a, "Value": v, "Color": COLOR_MAP[cname][1]})
            with grid[i % 2]:
                st.markdown(f"**{label}**")
                if not rows:
                    st.info("No comparable data.")
                    continue
                dfp = pd.DataFrame(rows)
                asc = rid in LEVERAGE_IDS  # leverage: lower better -> best at top
                dfp = dfp.sort_values("Value", ascending=not asc)
                fig = go.Figure(go.Bar(
                    x=dfp["Value"], y=dfp["Airline"], orientation="h",
                    marker_color=list(dfp["Color"]),
                ))
                vals = [r["Value"] for r in rows]
                med = statistics.median(vals)
                fig.add_vline(x=med, line_dash="dash", line_color="#6b7280",
                              annotation_text="median", annotation_position="top")
                fig.update_layout(height=max(220, 60 + 32 * len(dfp)),
                                  margin=dict(l=10, r=10, t=10, b=10),
                                  xaxis_title=("%" if sym == "%" else "x"))
                st.plotly_chart(fig, use_container_width=True)

# ----------------------------------------------------------------------------
# TAB 4 \u2014 Financial Statements (yfinance; from cached bundle)
# ----------------------------------------------------------------------------
with tab4:
    st.subheader("Financial Statements")
    fs_airline = st.selectbox("Select Airline", airlines, key="fs_airline")
    fs_period = st.radio("Period", ["Annual", "Quarterly"], horizontal=True, key="fs_period")
    fs_ticker = data[fs_airline]["ticker"]
    n_periods = 3 if fs_period == "Annual" else 4

    bundle = data[fs_airline].get("bundle", {})
    stmts = {k: bundle.get(k, pd.DataFrame()) for k in ["fin", "bs", "cf", "qfin", "qbs", "qcf"]}

    inc_df = stmts["fin"] if fs_period == "Annual" else stmts["qfin"]
    bal_df = stmts["bs"] if fs_period == "Annual" else stmts["qbs"]
    cf_df = stmts["cf"] if fs_period == "Annual" else stmts["qcf"]

    if inc_df.empty and bal_df.empty and cf_df.empty:
        st.error(f"No financial-statement data returned from Yahoo Finance for {fs_ticker}. "
                 "This is common for delisted or recently acquired carriers, or when Yahoo "
                 "rate-limits the request. Try again shortly or pick another airline.")
    else:
        xlsx_bytes = build_yf_excel(fs_ticker, stmts)
        st.download_button(
            f"\u2b07 Download {fs_airline} Financials to Excel",
            data=xlsx_bytes,
            file_name=f"{fs_ticker}_Financials_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="fs_download",
        )
        st.caption("Source: Yahoo Finance (yfinance). Values in millions of the airline's local "
                   "reporting currency (2dp) unless marked EPS or %. Rows absent from Yahoo's data "
                   "are skipped. " + ("Quarterly %\u0394 is period-over-period (seasonal), "
                                      "not year-on-year." if fs_period == "Quarterly" else ""))

        i_labels, i_cols, i_rows = prepare_yf_statement(inc_df, INCOME_YF, fs_period, n_periods,
                                                        revenue_df=inc_df)
        if i_labels:
            st.markdown(render_yf_html("Income Statement", i_labels, i_cols, i_rows, fs_period),
                        unsafe_allow_html=True)
        else:
            st.info("No income statement data available.")

        b_labels, b_cols, b_rows = prepare_yf_statement(bal_df, BALANCE_YF, fs_period, n_periods)
        if b_labels:
            extra = (_yf_series(bal_df, "Total Assets", b_cols),
                     _yf_series(bal_df, "Total Liabilities Net Minority Interest", b_cols),
                     _yf_series(bal_df, "Total Equity Gross Minority Interest", b_cols))
            st.markdown(render_yf_html("Balance Sheet", b_labels, b_cols, b_rows, fs_period,
                                       balance_extra=extra), unsafe_allow_html=True)
        else:
            st.info("No balance sheet data available.")

        c_labels, c_cols, c_rows = prepare_yf_statement(cf_df, CASHFLOW_YF, fs_period, n_periods,
                                                        revenue_df=inc_df)
        if c_labels:
            st.markdown(render_yf_html("Cash Flow Statement", c_labels, c_cols, c_rows, fs_period),
                        unsafe_allow_html=True)
        else:
            st.info("No cash flow data available.")


# ----------------------------------------------------------------------------
# TAB 5 \u2014 Fleet Composition (live Wikipedia scrape)
# ----------------------------------------------------------------------------
with tab5:
    st.subheader("Fleet Composition")
    st.caption(
        "Fleet data is scraped live from Wikipedia at the time of loading. Accuracy depends "
        "on how recently each airline's Wikipedia page was updated by its editors. For "
        "mission-critical decisions, verify against the airline's official investor disclosures."
    )

    if "fleet_cache" not in st.session_state:
        st.session_state.fleet_cache = {}
    fc = st.session_state.fleet_cache

    for name in airlines:
        ticker = data[name]["ticker"]
        st.markdown("---")
        hc1, hc2 = st.columns([5, 1])
        hc1.markdown(f"### {name} \u2014 Fleet Composition")
        if hc2.button("Refresh Fleet Data", key=f"fleet_refresh_{ticker}"):
            fc.pop(name, None)
            st.rerun()

        if name not in fc:
            cands = fleet_name_candidates(name, ticker)
            with st.spinner(f"Scraping fleet data for {name} ..."):
                try:
                    fc[name] = scrape_fleet(cands)
                except Exception:
                    fc[name] = {"ok": False,
                                "url": _wiki_url(cands[0]) if cands else "https://en.wikipedia.org/"}

        res = fc[name]
        if not res.get("ok"):
            st.warning(f"Could not locate a valid fleet table for {name} on Wikipedia. "
                       f"View the page directly: {res.get('url')}")
            continue

        df = res["df"]
        st.caption(f"Source: Wikipedia (scraped live) | Total Aircraft: {res['total']}")
        if res.get("flagged"):
            st.caption("\u26a0 Some rows had non-numeric counts that were set to 0 \u2014 "
                       "verify those against the source page.")
        st.dataframe(df, use_container_width=True, hide_index=True)

        fig = px.bar(df.sort_values("In Service"), x="In Service", y="Aircraft Type",
                     orientation="h")
        fig.update_layout(height=max(240, 40 + 28 * len(df)),
                          margin=dict(l=10, r=10, t=10, b=10))
        st.plotly_chart(fig, use_container_width=True)

# ----------------------------------------------------------------------------
# TAB 6 \u2014 Reverse DCF
# ----------------------------------------------------------------------------
with tab6:
    st.subheader("Reverse DCF")
    st.caption("Solves for the Year-0 unlevered FCF implied by each airline's current "
               "enterprise value, using a live-recalculating WACC.")

    G = 0.03  # terminal growth rate (assumed)
    rf = fetch_risk_free()

    if rf is None:
        st.error("Could not fetch current Treasury yield. Reverse DCF unavailable.")
    else:
        st.caption(f"Risk-free rate (10Y Treasury, ^TNX): {_pct(rf)} \u00b7 "
                   "Terminal growth rate (assumed, approximates long-run U.S. GDP growth): 3.0%")
        for name in airlines:
            ticker = data[name]["ticker"]
            st.divider()
            st.markdown(f"### {name} ({ticker}) \u2014 Reverse DCF Analysis")
            try:
                bundle = data[name].get("bundle", {})
                info = bundle.get("info") or {}
                recs = data[name]["records"]
                rec = recs[-1] if recs else None

                beta = info.get("beta")
                market_cap = info.get("marketCap")
                total_debt = rec["total_debt"] if rec else None
                cash = (rec["cash"] if rec else None) or 0.0
                interest = rec["interest_expense"] if rec else None
                M = extract_yf_financials(bundle)
                sti = (M.get("sti", {}).get(rec["fy"]) if rec else 0) or 0.0

                if beta is None:
                    st.info(f"Beta unavailable for {name} \u2014 Reverse DCF cannot be calculated.")
                    continue
                if market_cap is None:
                    st.info(f"Market cap unavailable for {name} \u2014 Reverse DCF cannot be calculated.")
                    continue
                if not total_debt or total_debt <= 0:
                    st.info(f"Insufficient debt data for {name} \u2014 Reverse DCF cannot be calculated.")
                    continue
                if interest is None:
                    st.info(f"Insufficient debt-cost data for {name} \u2014 Reverse DCF cannot be calculated.")
                    continue

                erp = st.slider(f"Equity Risk Premium \u2014 {name}", min_value=3.0, max_value=8.0,
                                value=5.0, step=0.1, format="%.1f%%", key=f"erp_{ticker}") / 100.0

                cost_of_equity = rf + beta * erp
                cost_of_debt = interest / total_debt
                tax_rate, tax_fallback = _effective_tax_rate(bundle)
                after_tax_cod = cost_of_debt * (1 - tax_rate)
                total_capital = market_cap + total_debt
                w_e = market_cap / total_capital
                w_d = total_debt / total_capital
                wacc = w_e * cost_of_equity + w_d * after_tax_cod

                if wacc <= G + 0.005:
                    st.warning(f"WACC is too close to or below the terminal growth rate for {name} "
                               "given current assumptions. Try increasing the Equity Risk Premium.")
                    continue

                # EV per dollar of Year-0 FCF. Matches the spec's derivation
                # EV = FCF0*(1+g)/(WACC-g)  =>  multiplier = (1+g)/(WACC-g).
                # (The spec's formula line read (1+wacc); to use that instead,
                #  replace (1 + G) with (1 + wacc) below.)
                multiplier = (1 + G) / (wacc - G)

                current_ev = market_cap + total_debt - cash - sti
                implied_fcf = current_ev / multiplier
                ttm_fcf, nq = _ttm_fcf(bundle)

                c1, c2, c3 = st.columns(3)
                c1.metric("Implied Required FCF", _money(implied_fcf))
                ttm_label = "Actual TTM FCF" if nq >= 4 else f"Actual FCF ({nq}Q)"
                if ttm_fcf is None:
                    c2.metric(ttm_label, "\u2014")
                else:
                    c2.metric(ttm_label, _money(ttm_fcf))
                if ttm_fcf is not None and implied_fcf not in (None, 0):
                    gap = ttm_fcf - implied_fcf
                    gap_pct = gap / abs(implied_fcf) * 100
                    c3.metric("Gap (Actual \u2212 Implied)", _money(gap), delta=f"{gap_pct:+.1f}%")
                else:
                    c3.metric("Gap (Actual \u2212 Implied)", "\u2014")

                if nq and nq < 4:
                    st.caption(f"TTM based on {nq} quarter(s) (limited data available).")

                if ttm_fcf is None:
                    st.info("Actual TTM FCF unavailable (no quarterly cash-flow data) \u2014 "
                            "showing implied requirement only.")
                elif ttm_fcf >= implied_fcf:
                    st.success(f"Current cash generation supports the stock price. {name}'s trailing "
                               f"FCF of {_money(ttm_fcf)} meets or exceeds the {_money(implied_fcf)} "
                               "required by current market pricing.")
                else:
                    gap_pct = (implied_fcf - ttm_fcf) / abs(implied_fcf) * 100
                    st.warning(f"Stock price implies cash generation beyond current run-rate. The "
                               f"market is pricing in {gap_pct:.1f}% more unlevered FCF than {name} "
                               "is currently generating (TTM basis).")

                with st.expander(f"View WACC & DCF Assumptions \u2014 {name}", expanded=False):
                    tax_note = " (fallback 21% statutory rate used)" if tax_fallback else ""
                    rows = [
                        ("MARKET INPUTS", ""),
                        ("Risk-Free Rate (10Y Treasury)", _pct(rf)),
                        ("Beta", f"{beta:.2f}"),
                        ("Equity Risk Premium", _pct(erp)),
                        ("\u2192 Cost of Equity", _pct(cost_of_equity)),
                        ("DEBT INPUTS", ""),
                        ("Interest Expense (latest FY)", _money(interest)),
                        ("Total Debt (incl. leases)", _money(total_debt)),
                        ("Cost of Debt", _pct(cost_of_debt)),
                        ("Effective Tax Rate", _pct(tax_rate) + tax_note),
                        ("\u2192 After-Tax Cost of Debt", _pct(after_tax_cod)),
                        ("CAPITAL STRUCTURE", ""),
                        ("Market Cap", _money(market_cap)),
                        ("Total Debt", _money(total_debt)),
                        ("Weight of Equity", f"{w_e * 100:.1f}%"),
                        ("Weight of Debt", f"{w_d * 100:.1f}%"),
                        ("\u2192 WACC", _pct(wacc)),
                        ("TERMINAL VALUE ASSUMPTIONS", ""),
                        ("Terminal Growth Rate", "3.0% (fixed assumption)"),
                        ("EV/FCF Multiplier", f"{multiplier:.1f}x"),
                        ("VALUATION BRIDGE", ""),
                        ("Current Market Cap", _money(market_cap)),
                        ("Plus Total Debt", _money(total_debt)),
                        ("Less Cash & STI", _money(cash + sti)),
                        ("= Current Enterprise Value", _money(current_ev)),
                        ("\u00f7 EV/FCF Multiplier", f"{multiplier:.1f}x"),
                        ("= Implied Required FCF", _money(implied_fcf)),
                    ]
                    html = ["<table style='width:100%;border-collapse:collapse;font-size:13px;'>"]
                    for label, val in rows:
                        if val == "":
                            html.append(
                                f"<tr><td colspan='2' style='background:#1e293b;color:#fff;"
                                f"font-weight:700;padding:5px 10px;letter-spacing:.04em;'>{label}</td></tr>")
                        else:
                            html.append(
                                f"<tr style='border-bottom:1px solid #f1f5f9;'>"
                                f"<td style='padding:4px 10px;'>{label}</td>"
                                f"<td style='padding:4px 10px;text-align:right;font-weight:600;'>{val}</td></tr>")
                    html.append("</table>")
                    st.markdown("".join(html), unsafe_allow_html=True)
                    st.caption(
                        "This is a simplified single-stage reverse DCF assuming constant terminal "
                        "growth from Year 0. It does not model a multi-year explicit growth ramp. "
                        "WACC recalculates live based on current market data (Treasury yield, beta, "
                        "market cap) each time this page loads, so figures will shift slightly between "
                        "visits even without company-specific news \u2014 this reflects real-time market "
                        "conditions, not a calculation error. Note: the implied figure is unlevered "
                        "while the TTM actual (OCF + CapEx) is a levered proxy, so the gap is indicative.")
            except Exception as exc:
                st.warning(f"Reverse DCF could not be completed for {name}: {exc}")

# ----------------------------------------------------------------------------
# TAB 7 \u2014 Fuel Analysis
# ----------------------------------------------------------------------------
with tab7:
    st.subheader("Fuel Analysis")

    # ---- Section 1: EIA jet fuel price chart ----
    st.markdown("#### 1 \u00b7 Jet Fuel Price")
    df_fuel = fetch_eia_jet_fuel()
    if df_fuel is None:
        st.error("Could not load EIA jet fuel price data. Check your connection and try again.")
    else:
        choice = st.selectbox("Timeframe",
                              ["Year to Date", "Last Month", "Last 12 Months", "Last 3 Years"],
                              index=2, key="fuel_tf")
        fdf = _filter_fuel_timeframe(df_fuel, choice).copy()
        if fdf.empty:
            st.info("No jet fuel data in the selected timeframe.")
        else:
            fdf["ma4"] = fdf["value"].rolling(4, min_periods=1).mean()
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=fdf["period"], y=fdf["value"], mode="lines",
                                     name="Weekly", line=dict(color="#2563eb")))
            fig.add_trace(go.Scatter(x=fdf["period"], y=fdf["ma4"], mode="lines",
                                     name="4-week avg", line=dict(color="#93c5fd", width=2)))
            fig.update_layout(title="U.S. Jet Fuel Price (Gulf Coast, $/gallon)", height=380,
                              margin=dict(l=10, r=10, t=40, b=10), yaxis_title="$/gallon")
            st.plotly_chart(fig, use_container_width=True)
            st.caption("Source: U.S. Energy Information Administration (EIA). "
                       "Weekly spot price, U.S. Gulf Coast.")

    st.divider()

    # ---- Section 2: fuel vs EBITDAR margin correlation ----
    st.markdown("#### 2 \u00b7 Fuel Price vs EBITDAR Margin")
    st.info("Correlation is calculated on the last 4 available quarters of data \u2014 the maximum "
            "reliably available from yfinance quarterly financials. Minimum 4 quarters required. "
            "Correlation is always calculated on trailing quarters regardless of the fuel chart "
            "timeframe selected above. With only 4 points the coefficient is directional, not "
            "statistically robust.")
    if df_fuel is None:
        st.warning("Fuel price data is unavailable, so fuel-vs-margin correlation cannot be computed.")
    else:
        for name in airlines:
            bundle = data[name].get("bundle", {})
            qm = _quarter_ebitdar_margins(bundle)
            xs, ys, labels = [], [], []
            if not qm.empty:
                for _, r in qm.iterrows():
                    fp = _avg_fuel_for_quarter(df_fuel, r["q_end"])
                    if fp is not None and pd.notna(fp) and pd.notna(r["margin"]):
                        xs.append(float(fp)); ys.append(float(r["margin"])); labels.append(r["label"])
            if len(xs) < 4:
                st.warning(f"Insufficient quarterly data for {name} to calculate fuel correlation. "
                           "Minimum 4 quarters required.")
                continue
            r_val, slope, intercept = _corr(xs, ys)
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=xs, y=ys, mode="markers+text", text=labels,
                                     textposition="top center", name="Quarters",
                                     marker=dict(size=11, color="#2563eb")))
            if slope is not None:
                lo, hi = min(xs), max(xs)
                fig.add_trace(go.Scatter(x=[lo, hi], y=[slope * lo + intercept, slope * hi + intercept],
                                         mode="lines", name="Trend", line=dict(color="#9ca3af", dash="dash")))
            fig.update_layout(title=f"{name} \u2014 Jet Fuel Price vs EBITDAR Margin (Last 4 Quarters)",
                              height=360, margin=dict(l=10, r=10, t=40, b=10),
                              xaxis_title="Avg quarterly fuel price ($/gal)", yaxis_title="EBITDAR margin (%)")
            st.plotly_chart(fig, use_container_width=True)
            if r_val is None:
                st.caption("Correlation could not be computed.")
            else:
                if r_val < -0.5:
                    col = GREEN_TEXT
                elif r_val <= -0.2:
                    col = AMBER_TEXT
                else:
                    col = RED_TEXT
                st.markdown(f"<span style='color:{col};font-weight:700;'>Correlation coefficient: "
                            f"{r_val:.2f}</span>", unsafe_allow_html=True)
            st.caption("A negative correlation indicates EBITDAR margin compresses as fuel prices "
                       "rise, as expected. Weak or positive correlation may indicate effective "
                       "hedging, fuel surcharge pass-through, or insufficient data.")

    st.divider()

    # ---- Section 3: fuel price sensitivity (U.S. carriers only) ----
    st.markdown("#### 3 \u00b7 Fuel Price Sensitivity")
    st.warning("Fuel sensitivity analysis is available for U.S. carriers only. This feature "
               "requires airline-specific fuel cost data from SEC EDGAR (XBRL tag: AirlineFuelCosts), "
               "which is only available for U.S.-listed carriers filing with the SEC. International "
               "airlines loaded in this session will not appear in this section.")

    us_names = [n for n in airlines
                if _is_us_carrier(data[n]["ticker"], data[n].get("bundle", {}).get("info") or {})]
    if not us_names:
        st.info("No U.S.-listed carriers loaded.")

    for name in us_names:
        ticker = data[name]["ticker"]
        st.markdown(f"**{name} ({ticker})**")
        cik = US_CARRIER_CIKS.get(ticker)
        if not cik:
            st.info(f"Fuel cost data not available in SEC EDGAR for {name}.")
            continue
        try:
            fuel_cost, opex = fetch_edgar_fuel(cik)
        except Exception:
            fuel_cost, opex = None, None
        if fuel_cost is None:
            st.info(f"Fuel cost data not available in SEC EDGAR for {name}.")
            continue

        rec = data[name]["records"][-1] if data[name]["records"] else None
        bundle = data[name].get("bundle", {})
        M = extract_yf_financials(bundle)
        fy = rec["fy"] if rec else None
        revenue = rec["revenue"] if rec else None
        interest = rec["interest_expense"] if rec else None
        ebitda = rec["ebitda"] if rec else None
        ebitdar = rec["ebitdar"] if rec else None
        oi = M.get("operating_income", {}).get(fy) if fy is not None else None
        if None in (revenue, ebitdar, oi) or revenue == 0:
            st.info(f"Baseline income data incomplete for {name} \u2014 sensitivity unavailable.")
            continue

        fuel_pct = (fuel_cost / opex * 100) if (opex and opex > 0) else None
        inc = st.slider(f"Fuel Price Increase \u2014 {name}", min_value=0, max_value=100,
                        value=10, step=5, format="+%d%%", key=f"fuel_sensitivity_{ticker}")

        impact = fuel_cost * inc / 100.0
        new_oi = oi - impact
        new_ebitdar = ebitdar - impact           # lease add-back unchanged by fuel
        base_margin = ebitdar / revenue * 100
        new_margin = new_ebitdar / revenue * 100

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Fuel Cost Impact", _money(impact),
                  help=(f"Fuel = {fuel_pct:.1f}% of operating expenses" if fuel_pct is not None else None))
        pct_change = (-impact / ebitdar * 100) if ebitdar else 0.0
        c2.metric("EBITDAR Change", _money(-impact), delta=f"{pct_change:.1f}%")
        c3.metric("New EBITDAR Margin", f"{new_margin:.1f}%",
                  help=f"Baseline {base_margin:.1f}%")

        if interest and interest > 0:
            cov_ic = absolute_rag(7, new_oi / interest)
            cov_er = absolute_rag(8, new_ebitdar / interest)
            order = {"red": 3, "amber": 2, "green": 1, None: 0}
            worst = max((cov_ic, cov_er), key=lambda c: order.get(c, 0))
            if worst == "green":
                c4.markdown(f"<div style='color:{GREEN_TEXT};font-weight:700;padding-top:12px;'>"
                            "\u2713 Covenants Clear</div>", unsafe_allow_html=True)
            elif worst == "amber":
                c4.markdown(f"<div style='color:{AMBER_TEXT};font-weight:700;padding-top:12px;'>"
                            "\u26a0 Approaching Threshold</div>", unsafe_allow_html=True)
            else:
                c4.markdown(f"<div style='color:{RED_TEXT};font-weight:700;padding-top:12px;'>"
                            "\u2717 Covenant Breach Risk</div>", unsafe_allow_html=True)
        else:
            c4.markdown("<div style='color:#6b7280;padding-top:12px;'>Coverage N/A</div>",
                        unsafe_allow_html=True)

        fig = go.Figure(go.Bar(x=["Baseline", f"+{inc}% fuel"], y=[base_margin, new_margin],
                               marker_color=["#2563eb", "#b91c1c"]))
        fig.update_layout(height=260, margin=dict(l=10, r=10, t=10, b=10),
                          yaxis_title="EBITDAR margin (%)")
        st.plotly_chart(fig, use_container_width=True)
        st.caption("Sensitivity based on most recent fiscal year fuel cost from SEC EDGAR. Assumes "
                   "fuel cost increase flows directly to operating expenses with no revenue offset "
                   "(pass-through, hedging, or demand elasticity effects not modeled). "
                   "Conservative estimate.")

# To run: streamlit run app.py
